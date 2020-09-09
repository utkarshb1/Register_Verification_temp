# from openpyxl import Workbook

# workbook = Workbook()
# sheet = workbook.active

# sheet["A1"] = "hello"
# sheet["B1"] = "world!"
# # 
# # workbook.save(filename="hello_world.gnumeric")
# workbook.save(filename="hello_world.xlsx")

# import json
# from openpyxl import load_workbook

# workbook = load_workbook(filename="reviews-sample.xlsx")
# sheet = workbook.active

# products = {}

# # Using the values_only because you want to return the cells' values
# for row in sheet.iter_rows(min_row=2,
#                            min_col=4,
#                            max_col=7,
#                            values_only=True):
#     product_id = row[0]
#     product = {
#         "parent": row[1],
#         "title": row[2],
#         "category": row[3]
#     }
#     products[product_id] = product

# Using json here to be able to format the output for displaying later
# print(json.dumps(products))
from openpyxl import load_workbook

# Start by opening the spreadsheet and selecting the main sheet
filename="hello_world.xlsx"
workbook = load_workbook(filename)
sheet = workbook.active

# Write what you want into a specific cell
sheet["C1"] = "writing ;)"

# Save the spreadsheet
workbook.save(filename=filename)